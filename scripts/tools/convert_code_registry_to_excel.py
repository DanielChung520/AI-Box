# 代碼功能說明: 代碼管制表轉Excel格式腳本
# 創建日期: 2025-01-27
# 創建人: Daniel Chung
# 最後修改日期: 2025-01-27

"""
代碼管制表轉Excel格式腳本

將 Markdown 格式的代碼管制表轉換為 Excel 格式，並添加"相關文件"欄位。

用法:
    python convert_code_registry_to_excel.py

輸入:
    docs/系统设计文档/代碼管制表.md

輸出:
    docs/代碼管制表.xlsx
"""

import re
import sys
from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("錯誤: 未安裝 openpyxl 庫")
    print("請運行: pip install openpyxl")
    sys.exit(1)

# 項目根目錄
PROJECT_ROOT = Path(__file__).parent

# 輸入文件路徑
INPUT_FILE = PROJECT_ROOT / "docs" / "系统设计文档" / "代碼管制表.md"

# 輸出文件路徑
OUTPUT_FILE = PROJECT_ROOT / "docs" / "代碼管制表.xlsx"


def parse_markdown_table(content: str) -> tuple[List[List[str]], List[str]]:
    """
    解析 Markdown 表格內容

    Args:
        content: Markdown 文件內容

    Returns:
        (數據行列表, 表頭列表)
    """
    lines = content.split("\n")
    header: Optional[List[str]] = None
    data_rows: List[List[str]] = []

    for line in lines:
        line_stripped = line.strip()

        # 檢查是否遇到統計信息標題（停止解析）
        if "## 統計信息" in line_stripped:
            break

        # 跳過空行和分隔行
        if (
            not line_stripped
            or line_stripped.startswith("---")
            or not line_stripped.startswith("|")
        ):
            continue

        # 解析表格行
        # 移除首尾的 |，然後按 | 分割
        parts = [part.strip() for part in line_stripped.split("|")[1:-1]]

        # 檢查是否為表頭（通常表頭在分隔行之前）
        if header is None and len(parts) >= 8:
            # 檢查是否包含"功能模組"等關鍵字
            if "功能模組" in parts[0] or "編號" in parts[1]:
                header = parts
                continue

        # 如果是數據行
        if header is not None and len(parts) >= 8:
            # 跳過分隔行（只包含 - 和空格的行）
            if all(re.match(r"^[\s-]*$", part) for part in parts):
                continue

            # 清理欄位內容（移除Markdown格式）
            cleaned_parts = []
            for part in parts:
                # 移除代碼反引號
                cleaned = re.sub(r"`([^`]+)`", r"\1", part)
                # 移除粗體標記
                cleaned = re.sub(r"\*\*([^*]+)\*\*", r"\1", cleaned)
                # 移除多餘空格
                cleaned = cleaned.strip()
                cleaned_parts.append(cleaned)
            data_rows.append(cleaned_parts)

    return data_rows, header or []


def create_excel_file(data_rows: List[List[str]], header: List[str]) -> None:
    """
    創建 Excel 文件

    Args:
        data_rows: 數據行列表
        header: 表頭列表
    """
    print(f"📝 創建 Excel 文件: {OUTPUT_FILE}")

    # 創建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "代碼管制表"

    # 定義新的表頭（添加"相關文件"欄位）
    new_header = [
        "功能模組",
        "編號",
        "名稱",
        "代碼",
        "代碼功能描述",
        "創建日期",
        "最後更新日期",
        "相關文件",  # 新添加的欄位
        "備註",
    ]

    # 寫入表頭
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(new_header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 寫入數據行
    print(f"📊 處理 {len(data_rows)} 條記錄...")
    for row_idx, row_data in enumerate(data_rows, start=2):
        if len(row_data) >= 8:
            # 原始欄位：功能模組、編號、名稱、代碼、代碼功能描述、創建日期、最後更新日期、備註
            module = row_data[0] if len(row_data) > 0 else ""
            number = row_data[1] if len(row_data) > 1 else ""
            name = row_data[2] if len(row_data) > 2 else ""
            code = row_data[3] if len(row_data) > 3 else ""
            description = row_data[4] if len(row_data) > 4 else ""
            created_date = row_data[5] if len(row_data) > 5 else ""
            last_updated = row_data[6] if len(row_data) > 6 else ""
            notes = row_data[7] if len(row_data) > 7 else ""

            # 新欄位順序：功能模組、編號、名稱、代碼、代碼功能描述、創建日期、最後更新日期、相關文件、備註
            ws.cell(row=row_idx, column=1, value=module)
            ws.cell(row=row_idx, column=2, value=number)
            ws.cell(row=row_idx, column=3, value=name)
            ws.cell(row=row_idx, column=4, value=code)
            ws.cell(row=row_idx, column=5, value=description)
            ws.cell(row=row_idx, column=6, value=created_date)
            ws.cell(row=row_idx, column=7, value=last_updated)
            ws.cell(row=row_idx, column=8, value="")  # 相關文件（空）
            ws.cell(row=row_idx, column=9, value=notes)

            # 設置對齊方式
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if row_idx % 100 == 0:
            print(f"⏳ 處理進度: {row_idx - 1}/{len(data_rows)}")

    # 自動調整列寬
    print("📏 調整列寬...")
    column_widths = {
        "A": 12,  # 功能模組
        "B": 12,  # 編號
        "C": 20,  # 名稱
        "D": 35,  # 代碼
        "E": 40,  # 代碼功能描述
        "F": 12,  # 創建日期
        "G": 12,  # 最後更新日期
        "H": 30,  # 相關文件
        "I": 50,  # 備註
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 凍結表頭
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(OUTPUT_FILE)
    print(f"✅ Excel 文件已生成: {OUTPUT_FILE}")
    print(f"📈 總共 {len(data_rows)} 條記錄")


def main() -> None:
    """主函數"""
    print("🔄 開始轉換代碼管制表為 Excel 格式...\n")

    # 檢查輸入文件是否存在
    if not INPUT_FILE.exists():
        print(f"❌ 錯誤: 輸入文件不存在: {INPUT_FILE}")
        sys.exit(1)

    # 讀取 Markdown 文件
    print(f"📖 讀取文件: {INPUT_FILE}")
    try:
        with open(INPUT_FILE, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        print(f"❌ 錯誤: 無法讀取文件: {e}")
        sys.exit(1)

    # 解析表格
    print("🔍 解析 Markdown 表格...")
    data_rows, header = parse_markdown_table(content)

    if not data_rows:
        print("❌ 錯誤: 未找到數據行")
        sys.exit(1)

    print(f"✅ 解析完成，找到 {len(data_rows)} 條記錄\n")

    # 確保輸出目錄存在
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # 創建 Excel 文件
    create_excel_file(data_rows, header)

    print("\n✨ 轉換完成！")


if __name__ == "__main__":
    main()
