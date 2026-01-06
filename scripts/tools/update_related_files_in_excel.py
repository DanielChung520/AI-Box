# 代碼功能說明: 更新Excel文件中相關文件欄位
# 創建日期: 2025-01-27
# 創建人: Daniel Chung
# 最後修改日期: 2025-01-27

"""
更新Excel文件中相關文件欄位

掃描 docs/系统设计文档 目錄下的所有文件，檢查每個代碼是否在文檔中被提及，
如果找到，填入專案相對路徑/檔名到"相關文件"欄位。

用法:
    python update_related_files_in_excel.py

輸入:
    docs/代碼管制表.xlsx

輸出:
    docs/代碼管制表.xlsx (更新後的檔案)
"""

import re
import sys
from pathlib import Path
from typing import List, Set

try:
    from openpyxl import load_workbook
except ImportError:
    print("錯誤: 未安裝 openpyxl 庫")
    print("請運行: pip install openpyxl")
    sys.exit(1)

# 項目根目錄
PROJECT_ROOT = Path(__file__).parent

# Excel文件路徑
EXCEL_FILE = PROJECT_ROOT / "docs" / "代碼管制表.xlsx"

# 文檔目錄
DOCS_DIR = PROJECT_ROOT / "docs" / "系统设计文档"

# 欄位索引（從1開始）
COL_CODE = 4  # 代碼欄位（第4列）
COL_RELATED_FILES = 8  # 相關文件欄位（第8列）


def normalize_code_path(code_path: str) -> str:
    """
    標準化代碼路徑，用於匹配

    Args:
        code_path: 代碼路徑（可能包含反引號等格式）

    Returns:
        標準化後的路徑
    """
    # 移除反引號和空格
    normalized = code_path.strip().strip("`").strip()
    # 移除開頭的斜線（如果有）
    if normalized.startswith("/"):
        normalized = normalized[1:]
    return normalized


def extract_code_basename(code_path: str) -> str:
    """
    提取代碼路徑的檔案名（不含擴展名和路徑）

    Args:
        code_path: 代碼路徑

    Returns:
        檔案名（不含擴展名）
    """
    normalized = normalize_code_path(code_path)
    # 提取檔案名
    if "/" in normalized:
        filename = normalized.split("/")[-1]
    else:
        filename = normalized
    # 移除擴展名
    if "." in filename:
        filename = filename.rsplit(".", 1)[0]
    return filename


def find_markdown_files(docs_dir: Path) -> List[Path]:
    """
    查找所有 Markdown 文件

    Args:
        docs_dir: 文檔目錄

    Returns:
        Markdown 文件列表
    """
    md_files = []
    for md_file in docs_dir.rglob("*.md"):
        md_files.append(md_file)
    return sorted(md_files)


def check_code_in_file(code_path: str, file_path: Path) -> bool:
    """
    檢查代碼路徑是否在文件中被提及

    Args:
        code_path: 代碼路徑
        file_path: 文件路徑

    Returns:
        是否找到
    """
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return False

    # 標準化代碼路徑
    normalized_code = normalize_code_path(code_path)

    # 提取檔案名（不含擴展名）
    code_basename = extract_code_basename(code_path)

    # 檢查方式1：完整路徑匹配（考慮可能包含反引號）
    patterns = [
        normalized_code,  # 完整路徑
        f"`{normalized_code}`",  # 反引號包圍
        code_path,  # 原始路徑
    ]

    # 檢查方式2：檔案名匹配（更寬鬆的匹配）
    # 檢查是否有模組路徑或檔案名
    basename_patterns = [
        code_basename,
        f"`{code_basename}`",
    ]

    # 如果檔案名是 __init__，則跳過（太常見）
    if code_basename == "__init__":
        # 只檢查完整路徑
        for pattern in patterns:
            if pattern in content:
                return True
    else:
        # 檢查完整路徑
        for pattern in patterns:
            if pattern in content:
                return True

        # 檢查檔案名（更寬鬆）
        # 確保檔案名出現在代碼相關的上下文中（例如在路徑中或作為檔案名提及）
        for pattern in basename_patterns:
            if pattern in content:
                # 進一步驗證：檢查是否在路徑上下文中
                # 使用正則表達式匹配更精確的上下文
                regex_pattern = rf"[\w/\.]*{re.escape(code_basename)}[\w\.]*"
                if re.search(regex_pattern, content):
                    return True

    return False


def get_relative_path(file_path: Path) -> str:
    """
    獲取文件相對於項目根目錄的路徑

    Args:
        file_path: 文件路徑

    Returns:
        相對路徑字符串
    """
    try:
        relative = file_path.relative_to(PROJECT_ROOT)
        return str(relative).replace("\\", "/")
    except ValueError:
        return str(file_path)


def update_excel_related_files() -> None:
    """更新Excel文件中的相關文件欄位"""
    print(f"📖 讀取 Excel 文件: {EXCEL_FILE}")

    # 檢查文件是否存在
    if not EXCEL_FILE.exists():
        print(f"❌ 錯誤: Excel 文件不存在: {EXCEL_FILE}")
        sys.exit(1)

    # 加載工作簿
    workbook = load_workbook(EXCEL_FILE)
    worksheet = workbook.active

    # 查找所有 Markdown 文件
    print(f"🔍 掃描文檔目錄: {DOCS_DIR}")
    md_files = find_markdown_files(DOCS_DIR)
    print(f"✅ 找到 {len(md_files)} 個 Markdown 文件\n")

    # 統計信息
    total_rows = 0
    updated_rows = 0

    # 遍歷每一行（跳過表頭，從第2行開始）
    print("🔎 檢查代碼是否在文檔中被提及...\n")
    for row_idx in range(2, worksheet.max_row + 1):
        code_cell = worksheet.cell(row=row_idx, column=COL_CODE)
        related_files_cell = worksheet.cell(row=row_idx, column=COL_RELATED_FILES)

        code_path = code_cell.value
        if not code_path or not isinstance(code_path, str):
            continue

        total_rows += 1
        found_files: Set[str] = set()

        # 檢查每個 Markdown 文件
        for md_file in md_files:
            if check_code_in_file(code_path, md_file):
                relative_path = get_relative_path(md_file)
                found_files.add(relative_path)

        # 如果找到相關文件，更新欄位
        if found_files:
            # 如果欄位已有內容，追加（用分號分隔）
            existing_value = related_files_cell.value or ""
            if existing_value and isinstance(existing_value, str):
                existing_files = set(f.strip() for f in existing_value.split(";") if f.strip())
                found_files.update(existing_files)

            # 排序並用分號分隔
            new_value = "; ".join(sorted(found_files))
            related_files_cell.value = new_value
            updated_rows += 1

            if total_rows % 100 == 0:
                print(f"⏳ 處理進度: {total_rows} 行，已更新 {updated_rows} 行")

    # 保存文件
    print("\n💾 保存 Excel 文件...")
    workbook.save(EXCEL_FILE)
    print(f"✅ Excel 文件已更新: {EXCEL_FILE}")
    print(f"📊 總共處理 {total_rows} 行，更新了 {updated_rows} 行的相關文件欄位")


def main() -> None:
    """主函數"""
    print("🔄 開始更新Excel文件中的相關文件欄位...\n")

    # 檢查文檔目錄是否存在
    if not DOCS_DIR.exists():
        print(f"❌ 錯誤: 文檔目錄不存在: {DOCS_DIR}")
        sys.exit(1)

    # 更新Excel文件
    update_excel_related_files()

    print("\n✨ 更新完成！")


if __name__ == "__main__":
    main()
