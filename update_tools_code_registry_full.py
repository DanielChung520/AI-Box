# 代碼功能說明: 更新工具開發計劃相關代碼到代碼管制表（完整版）
# 創建日期: 2025-12-30
# 創建人: Daniel Chung
# 最後修改日期: 2025-12-30

"""
更新工具開發計劃相關代碼到代碼管制表（完整版）

根據工具開發計劃管控表中的文件清單，更新代碼管制表.xlsx：
1. 更新所有相關欄位：功能模組、編號、名稱、代碼、代碼功能描述、創建日期、最後更新日期、開發難度等級/工時估算
2. 添加或更新所有工具相關的代碼文件
3. 包含文件、開發、測試等合計工時

用法:
    python update_tools_code_registry_full.py
"""

import sys
from pathlib import Path
from typing import Dict, Optional, TypedDict

try:
    from openpyxl import load_workbook
except ImportError:
    print("錯誤: 未安裝 openpyxl 庫")
    print("請運行: pip install openpyxl")
    sys.exit(1)


class FileInfo(TypedDict):
    """文件信息類型定義"""

    category: str
    difficulty: str
    hours: float
    description: str  # 新增：代碼功能描述


def get_module_name(code_path: str) -> str:
    """根據代碼路徑獲取功能模組名稱"""
    if code_path.startswith("tools/"):
        return "工具組"
    elif code_path.startswith("tests/tools/"):
        return "工具組-測試"
    elif code_path.startswith("docs/"):
        return "工具組-文檔"
    elif code_path.startswith("scripts/"):
        return "工具組-腳本"
    else:
        return "工具組"


def get_file_name(code_path: str) -> str:
    """從代碼路徑提取文件名（不含路徑和擴展名）"""
    path = Path(code_path)
    name = path.stem  # 不含擴展名

    # 處理特殊文件名
    if name == "__init__":
        # 從路徑中提取模組名
        parts = path.parts
        if len(parts) > 1:
            return f"{parts[-2]}模組初始化"
        return "模組初始化"

    # 將下劃線轉換為空格，並轉換為中文友好的名稱
    name = name.replace("_", " ").replace("-", " ")

    # 常見文件名映射
    name_mapping = {
        "base": "基礎類",
        "registry": "註冊表",
        "errors": "錯誤定義",
        "validator": "驗證器",
        "cache": "緩存工具",
        "datetime tool": "日期時間工具",
        "formatter": "格式化工具",
        "calculator": "計算工具",
        "weather tool": "天氣工具",
        "forecast tool": "天氣預報工具",
        "ip location": "IP定位工具",
        "geocoding": "地理編碼工具",
        "distance": "距離計算工具",
        "timezone": "時區查詢工具",
        "length": "長度單位轉換",
        "weight": "重量單位轉換",
        "temperature": "溫度單位轉換",
        "currency": "貨幣轉換",
        "volume": "體積單位轉換",
        "area": "面積單位轉換",
        "math calculator": "數學計算工具",
        "statistics": "統計計算工具",
        "cleaner": "文本清理工具",
        "converter": "文本轉換工具",
        "summarizer": "文本摘要工具",
        "smart time service": "智能時間服務",
    }

    return name_mapping.get(name.lower(), name)


def get_description(code_path: str, category: str, file_info: FileInfo) -> str:
    """生成代碼功能描述"""
    if file_info.get("description"):
        return file_info["description"]

    # 根據類別和文件名生成描述
    path = Path(code_path)
    name = path.stem

    descriptions = {
        "基礎設施": {
            "base": "工具基類，定義所有工具的統一接口和基礎功能",
            "registry": "工具註冊表，管理工具的註冊、查找和調用",
            "errors": "工具相關錯誤定義",
            "validator": "參數驗證工具",
            "cache": "緩存工具，用於外部API調用的緩存機制",
        },
        "時間工具": {
            "datetime_tool": "日期時間工具，獲取當前日期時間，支持時區轉換和配置管理",
            "formatter": "日期格式化工具，支持多種格式和語言環境",
            "calculator": "日期計算工具，日期差值計算、加減運算、工作日計算",
            "smart_time_service": "智能時間服務，提供高精度時間和緩存機制",
        },
        "天氣工具": {
            "weather_tool": "天氣查詢工具，根據地理位置獲取當前天氣信息",
            "forecast_tool": "天氣預報工具，獲取未來幾天的天氣預報",
            "base": "天氣提供商基類",
            "openweathermap": "OpenWeatherMap天氣API提供商實現",
        },
        "地理位置工具": {
            "ip_location": "IP地址定位工具，根據IP地址獲取地理位置信息",
            "geocoding": "地理編碼工具，地址與經緯度之間的轉換",
            "distance": "距離計算工具，計算兩個地理位置之間的距離",
            "timezone": "時區查詢工具，根據地理位置獲取時區信息",
        },
        "單位轉換工具": {
            "length": "長度單位轉換工具",
            "weight": "重量單位轉換工具",
            "temperature": "溫度單位轉換工具",
            "currency": "貨幣轉換工具，支持實時匯率API",
            "volume": "體積單位轉換工具",
            "area": "面積單位轉換工具",
        },
        "計算工具": {
            "math_calculator": "數學計算工具，支持基本運算和科學計算",
            "statistics": "統計計算工具，平均值、中位數、標準差等統計功能",
        },
        "文本處理工具": {
            "formatter": "文本格式化工具，大小寫轉換、首字母大寫等",
            "cleaner": "文本清理工具，去除空白、特殊字符等",
            "converter": "文本轉換工具，Markdown、HTML、純文本互轉",
            "summarizer": "文本摘要工具，提取關鍵信息",
        },
        "測試": {
            "test": "單元測試",
            "integration": "集成測試",
        },
        "文檔": {
            "工具API文档": "工具API完整文檔，包含所有工具的API說明",
            "工具使用指南": "工具使用指南，包含快速開始、常見場景、最佳實踐",
        },
    }

    category_desc = descriptions.get(category, {})
    desc = category_desc.get(name, "")

    if not desc:
        # 默認描述
        file_name = get_file_name(code_path)
        desc = f"{file_name}相關功能實現"

    return desc


# 項目根目錄
PROJECT_ROOT = Path(__file__).parent

# Excel文件路徑
EXCEL_FILE = PROJECT_ROOT / "docs" / "代碼管制表.xlsx"

# 創建日期和更新日期
CREATED_DATE = "2025-12-30"
UPDATED_DATE = "2025-12-30"

# 工具開發計劃文件清單（從管控表中提取，添加描述）
TOOLS_FILES: Dict[str, FileInfo] = {
    # 基礎設施（6 個文件）
    "tools/base.py": {
        "category": "基礎設施",
        "difficulty": "中",
        "hours": 4,
        "description": "工具基類，定義所有工具的統一接口和基礎功能",
    },
    "tools/registry.py": {
        "category": "基礎設施",
        "difficulty": "中",
        "hours": 4,
        "description": "工具註冊表，管理工具的註冊、查找和調用",
    },
    "tools/utils/errors.py": {
        "category": "基礎設施",
        "difficulty": "低",
        "hours": 2,
        "description": "工具相關錯誤定義",
    },
    "tools/utils/validator.py": {
        "category": "基礎設施",
        "difficulty": "低",
        "hours": 2,
        "description": "參數驗證工具",
    },
    "tools/utils/cache.py": {
        "category": "基礎設施",
        "difficulty": "中",
        "hours": 3,
        "description": "緩存工具，用於外部API調用的緩存機制",
    },
    "tools/utils/__init__.py": {
        "category": "基礎設施",
        "difficulty": "低",
        "hours": 0.5,
        "description": "工具工具模組初始化",
    },
    # 時間工具（4 個文件）
    "tools/time/datetime_tool.py": {
        "category": "時間工具",
        "difficulty": "高",
        "hours": 6,
        "description": "日期時間工具，獲取當前日期時間，支持時區轉換和配置管理",
    },
    "tools/time/formatter.py": {
        "category": "時間工具",
        "difficulty": "中",
        "hours": 4,
        "description": "日期格式化工具，支持多種格式和語言環境",
    },
    "tools/time/calculator.py": {
        "category": "時間工具",
        "difficulty": "中",
        "hours": 4,
        "description": "日期計算工具，日期差值計算、加減運算、工作日計算",
    },
    "tools/time/__init__.py": {
        "category": "時間工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "時間工具模組初始化",
    },
    # 天氣工具（5 個文件）
    "tools/weather/weather_tool.py": {
        "category": "天氣工具",
        "difficulty": "中",
        "hours": 4,
        "description": "天氣查詢工具，根據地理位置獲取當前天氣信息",
    },
    "tools/weather/providers/base.py": {
        "category": "天氣工具",
        "difficulty": "中",
        "hours": 3,
        "description": "天氣提供商基類",
    },
    "tools/weather/providers/openweathermap.py": {
        "category": "天氣工具",
        "difficulty": "中",
        "hours": 4,
        "description": "OpenWeatherMap天氣API提供商實現",
    },
    "tools/weather/providers/__init__.py": {
        "category": "天氣工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "天氣提供商模組初始化",
    },
    "tools/weather/__init__.py": {
        "category": "天氣工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "天氣工具模組初始化",
    },
    # 地理位置工具（4 個文件）
    "tools/location/ip_location.py": {
        "category": "地理位置工具",
        "difficulty": "中",
        "hours": 4,
        "description": "IP地址定位工具，根據IP地址獲取地理位置信息",
    },
    "tools/location/geocoding.py": {
        "category": "地理位置工具",
        "difficulty": "中",
        "hours": 4,
        "description": "地理編碼工具，地址與經緯度之間的轉換",
    },
    "tools/location/distance.py": {
        "category": "地理位置工具",
        "difficulty": "中",
        "hours": 4,
        "description": "距離計算工具，計算兩個地理位置之間的距離",
    },
    "tools/location/__init__.py": {
        "category": "地理位置工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "地理位置工具模組初始化",
    },
    # 擴展功能（Phase 2）
    "tools/location/timezone.py": {
        "category": "地理位置工具",
        "difficulty": "中",
        "hours": 3,
        "description": "時區查詢工具，根據地理位置獲取時區信息",
    },
    "tools/weather/forecast_tool.py": {
        "category": "天氣工具",
        "difficulty": "中",
        "hours": 4,
        "description": "天氣預報工具，獲取未來幾天的天氣預報",
    },
    # 單位轉換工具（7 個文件）
    "tools/conversion/__init__.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "單位轉換工具模組初始化",
    },
    "tools/conversion/length.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 2,
        "description": "長度單位轉換工具",
    },
    "tools/conversion/weight.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 2,
        "description": "重量單位轉換工具",
    },
    "tools/conversion/temperature.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 2,
        "description": "溫度單位轉換工具",
    },
    "tools/conversion/currency.py": {
        "category": "單位轉換工具",
        "difficulty": "中",
        "hours": 4,
        "description": "貨幣轉換工具，支持實時匯率API",
    },
    "tools/conversion/volume.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 2,
        "description": "體積單位轉換工具",
    },
    "tools/conversion/area.py": {
        "category": "單位轉換工具",
        "difficulty": "低",
        "hours": 2,
        "description": "面積單位轉換工具",
    },
    # 計算工具（3 個文件）
    "tools/calculator/__init__.py": {
        "category": "計算工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "計算工具模組初始化",
    },
    "tools/calculator/math_calculator.py": {
        "category": "計算工具",
        "difficulty": "中",
        "hours": 4,
        "description": "數學計算工具，支持基本運算和科學計算",
    },
    "tools/calculator/statistics.py": {
        "category": "計算工具",
        "difficulty": "中",
        "hours": 4,
        "description": "統計計算工具，平均值、中位數、標準差等統計功能",
    },
    # 文本處理工具（5 個文件）
    "tools/text/__init__.py": {
        "category": "文本處理工具",
        "difficulty": "低",
        "hours": 0.5,
        "description": "文本處理工具模組初始化",
    },
    "tools/text/formatter.py": {
        "category": "文本處理工具",
        "difficulty": "低",
        "hours": 2,
        "description": "文本格式化工具，大小寫轉換、首字母大寫等",
    },
    "tools/text/cleaner.py": {
        "category": "文本處理工具",
        "difficulty": "低",
        "hours": 2,
        "description": "文本清理工具，去除空白、特殊字符等",
    },
    "tools/text/converter.py": {
        "category": "文本處理工具",
        "difficulty": "中",
        "hours": 3,
        "description": "文本轉換工具，Markdown、HTML、純文本互轉",
    },
    "tools/text/summarizer.py": {
        "category": "文本處理工具",
        "difficulty": "中",
        "hours": 3,
        "description": "文本摘要工具，提取關鍵信息",
    },
    # 工具組初始化
    "tools/__init__.py": {
        "category": "工具組初始化",
        "difficulty": "低",
        "hours": 1,
        "description": "工具組模組初始化，註冊所有工具",
    },
    # 配置腳本
    "scripts/init_tools_datetime_config.py": {
        "category": "配置腳本",
        "difficulty": "低",
        "hours": 2,
        "description": "工具日期時間配置初始化腳本",
    },
    # 時間服務（smart_time_service）
    "tools/time/smart_time_service.py": {
        "category": "時間工具",
        "difficulty": "高",
        "hours": 5,
        "description": "智能時間服務，提供高精度時間和緩存機制",
    },
}

# 測試文件（30 個文件）- 簡化描述
TEST_FILES: Dict[str, FileInfo] = {
    # 單元測試（27 個文件）
    "tests/tools/__init__.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 0.5,
        "description": "測試模組初始化",
    },
    "tests/tools/conftest.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1,
        "description": "測試配置文件",
    },
    "tests/tools/test_base.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "工具基類單元測試",
    },
    "tests/tools/test_registry.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "工具註冊表單元測試",
    },
    "tests/tools/test_time_datetime.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "日期時間工具單元測試",
    },
    "tests/tools/test_time_formatter.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "日期格式化工具單元測試",
    },
    "tests/tools/test_time_calculator.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "日期計算工具單元測試",
    },
    "tests/tools/test_time_smart_time_service.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "智能時間服務單元測試",
    },
    "tests/tools/test_weather.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "天氣工具單元測試",
    },
    "tests/tools/test_weather_forecast.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "天氣預報工具單元測試",
    },
    "tests/tools/test_location_ip.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "IP定位工具單元測試",
    },
    "tests/tools/test_location_geocoding.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "地理編碼工具單元測試",
    },
    "tests/tools/test_location_distance.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "距離計算工具單元測試",
    },
    "tests/tools/test_location_timezone.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "時區查詢工具單元測試",
    },
    "tests/tools/test_conversion_length.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "長度單位轉換工具單元測試",
    },
    "tests/tools/test_conversion_weight.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "重量單位轉換工具單元測試",
    },
    "tests/tools/test_conversion_currency.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "貨幣轉換工具單元測試",
    },
    "tests/tools/test_conversion_temperature.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "溫度單位轉換工具單元測試",
    },
    "tests/tools/test_conversion_volume.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "體積單位轉換工具單元測試",
    },
    "tests/tools/test_conversion_area.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "面積單位轉換工具單元測試",
    },
    "tests/tools/test_calculator_math.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "數學計算工具單元測試",
    },
    "tests/tools/test_calculator_statistics.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "統計計算工具單元測試",
    },
    "tests/tools/test_text_formatter.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "文本格式化工具單元測試",
    },
    "tests/tools/test_text_cleaner.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "文本清理工具單元測試",
    },
    "tests/tools/test_text_converter.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "文本轉換工具單元測試",
    },
    "tests/tools/test_text_summarizer.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "文本摘要工具單元測試",
    },
    "tests/tools/test_utils_cache.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "緩存工具單元測試",
    },
    "tests/tools/test_utils_validator.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 2,
        "description": "驗證器單元測試",
    },
    "tests/tools/test_utils_errors.py": {
        "category": "測試",
        "difficulty": "低",
        "hours": 1.5,
        "description": "錯誤定義單元測試",
    },
    # 集成測試（3 個文件）
    "tests/tools/test_integration_registry.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 3,
        "description": "工具註冊表集成測試",
    },
    "tests/tools/test_integration_config.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 3,
        "description": "配置服務集成測試",
    },
    "tests/tools/test_integration_apis.py": {
        "category": "測試",
        "difficulty": "中",
        "hours": 3,
        "description": "外部API集成測試",
    },
}

# 文檔文件
DOC_FILES: Dict[str, FileInfo] = {
    "docs/系统设计文档/tools/工具API文档.md": {
        "category": "文檔",
        "difficulty": "中",
        "hours": 4,
        "description": "工具API完整文檔，包含所有工具的API說明",
    },
    "docs/系统设计文档/tools/工具使用指南.md": {
        "category": "文檔",
        "difficulty": "中",
        "hours": 4,
        "description": "工具使用指南，包含快速開始、常見場景、最佳實踐",
    },
}


def find_column_by_header(worksheet, header_text: str) -> Optional[int]:
    """根據表頭文本查找列號"""
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(1, col).value
        if cell_value and header_text in str(cell_value):
            return col
    return None


def find_existing_row(worksheet, code_path: str, code_col: int) -> Optional[int]:
    """查找代碼路徑是否已存在"""
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row, code_col).value
        if cell_value and str(cell_value).strip() == code_path.strip():
            return row
    return None


def generate_number(module: str, existing_numbers: Dict[str, int]) -> str:
    """生成編號"""
    # 簡化編號生成：使用模組前綴 + 序號
    prefix_map = {
        "工具組": "TOOL",
        "工具組-測試": "TOOL-TEST",
        "工具組-文檔": "TOOL-DOC",
        "工具組-腳本": "TOOL-SCRIPT",
    }

    prefix = prefix_map.get(module, "TOOL")
    if module not in existing_numbers:
        existing_numbers[module] = 0

    existing_numbers[module] += 1
    return f"{prefix}-{existing_numbers[module]:03d}"


def format_difficulty_hours(difficulty: str, hours: float) -> str:
    """格式化難度等級和工時"""
    return f"{difficulty} ({hours:.1f}小時)"


def update_excel() -> None:
    """更新Excel文件"""
    print(f"📖 讀取 Excel 文件: {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        print(f"❌ 錯誤: Excel 文件不存在: {EXCEL_FILE}")
        sys.exit(1)

    # 加載工作簿
    workbook = load_workbook(EXCEL_FILE)
    worksheet = workbook.active

    # 查找各欄位列號
    module_col = find_column_by_header(worksheet, "功能模組")
    number_col = find_column_by_header(worksheet, "編號")
    name_col = find_column_by_header(worksheet, "名稱")
    code_col = find_column_by_header(worksheet, "代碼")
    desc_col = find_column_by_header(worksheet, "代碼功能描述")
    created_col = find_column_by_header(worksheet, "創建日期")
    updated_col = find_column_by_header(worksheet, "最後更新日期")
    related_col = find_column_by_header(worksheet, "相關文件")
    difficulty_col = find_column_by_header(worksheet, "開發難度")

    if not code_col:
        print("❌ 錯誤: 找不到代碼欄位")
        sys.exit(1)

    print("✅ 找到欄位:")
    print(f"  - 功能模組: 第 {module_col} 列" if module_col else "  - 功能模組: 未找到")
    print(f"  - 編號: 第 {number_col} 列" if number_col else "  - 編號: 未找到")
    print(f"  - 名稱: 第 {name_col} 列" if name_col else "  - 名稱: 未找到")
    print(f"  - 代碼: 第 {code_col} 列")
    print(f"  - 代碼功能描述: 第 {desc_col} 列" if desc_col else "  - 代碼功能描述: 未找到")
    print(f"  - 開發難度: 第 {difficulty_col} 列" if difficulty_col else "  - 開發難度: 未找到")

    # 合併所有文件
    all_files = {**TOOLS_FILES, **TEST_FILES, **DOC_FILES}

    # 統計信息
    updated_count = 0
    added_count = 0
    existing_numbers: Dict[str, int] = {}

    # 更新或添加文件
    print(f"\n📝 開始更新 {len(all_files)} 個文件...\n")

    for code_path, file_info in all_files.items():
        # 查找是否已存在
        existing_row = find_existing_row(worksheet, code_path, code_col)

        difficulty = file_info["difficulty"]
        hours = file_info["hours"]
        category = file_info["category"]
        description = get_description(code_path, category, file_info)
        module = get_module_name(code_path)
        name = get_file_name(code_path)

        if existing_row:
            # 更新現有行
            if module_col:
                worksheet.cell(existing_row, module_col).value = module
            if number_col:
                # 如果沒有編號，生成一個
                existing_number = worksheet.cell(existing_row, number_col).value
                if not existing_number:
                    number = generate_number(module, existing_numbers)
                    worksheet.cell(existing_row, number_col).value = number
            if name_col:
                worksheet.cell(existing_row, name_col).value = name
            if desc_col:
                worksheet.cell(existing_row, desc_col).value = description
            if created_col and not worksheet.cell(existing_row, created_col).value:
                worksheet.cell(existing_row, created_col).value = CREATED_DATE
            if updated_col:
                worksheet.cell(existing_row, updated_col).value = UPDATED_DATE
            if difficulty_col:
                worksheet.cell(existing_row, difficulty_col).value = format_difficulty_hours(
                    difficulty, hours
                )
            if related_col:
                related_value = worksheet.cell(existing_row, related_col).value or ""
                if "工具開發計劃管控表" not in str(related_value):
                    new_related = (
                        f"{related_value}; docs/系统设计文档/tools/工具開發計劃管控表.md"
                        if related_value
                        else "docs/系统设计文档/tools/工具開發計劃管控表.md"
                    )
                    worksheet.cell(existing_row, related_col).value = new_related

            updated_count += 1
            print(f"  ✅ 更新: {code_path}")
        else:
            # 添加新行
            new_row = worksheet.max_row + 1

            # 生成編號
            number = generate_number(module, existing_numbers) if number_col else ""

            # 填寫各欄位
            if module_col:
                worksheet.cell(new_row, module_col).value = module
            if number_col:
                worksheet.cell(new_row, number_col).value = number
            if name_col:
                worksheet.cell(new_row, name_col).value = name
            worksheet.cell(new_row, code_col).value = code_path
            if desc_col:
                worksheet.cell(new_row, desc_col).value = description
            if created_col:
                worksheet.cell(new_row, created_col).value = CREATED_DATE
            if updated_col:
                worksheet.cell(new_row, updated_col).value = UPDATED_DATE
            if related_col:
                worksheet.cell(new_row, related_col).value = "docs/系统设计文档/tools/工具開發計劃管控表.md"
            if difficulty_col:
                worksheet.cell(new_row, difficulty_col).value = format_difficulty_hours(
                    difficulty, hours
                )

            added_count += 1
            print(f"  ➕ 新增: {code_path}")

    # 保存文件
    print("\n💾 保存 Excel 文件...")
    workbook.save(EXCEL_FILE)
    print(f"✅ Excel 文件已更新: {EXCEL_FILE}")
    print("\n📊 統計:")
    print(f"  - 更新: {updated_count} 個文件")
    print(f"  - 新增: {added_count} 個文件")
    print(f"  - 總計: {updated_count + added_count} 個文件")

    # 計算總工時
    total_hours = sum(info["hours"] for info in all_files.values())
    print(f"  - 總工時: {total_hours:.1f} 小時")


def main() -> None:
    """主函數"""
    try:
        update_excel()
        print("\n✨ 更新完成！")
    except Exception as e:
        print(f"\n❌ 錯誤: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
