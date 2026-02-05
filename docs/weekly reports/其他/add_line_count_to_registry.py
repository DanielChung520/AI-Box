# 代碼功能說明: 代碼管制表新增「代碼行數」欄位並盤點
# 創建日期: 2026-02-04
# 創建人: Daniel Chung
# 最後修改日期: 2026-02-04

"""
讀取代碼管制表 MD，為每個代碼路徑統計行數，插入「代碼行數」欄位，並輸出更新後的 MD 與 Excel。
"""

import re
import shutil
import subprocess
import sys
from pathlib import Path


def count_lines(file_path: Path) -> str:
    """統計檔案行數，若檔案不存在或非文字檔則回傳 '-'。"""
    if not file_path.exists():
        return "-"
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return str(sum(1 for _ in f))
    except Exception:
        return "-"


def extract_path(cell: str) -> str:
    """從表格儲存格取出路徑（去掉反引號）。"""
    s = cell.strip()
    if s.startswith("`") and s.endswith("`"):
        return s[1:-1].strip()
    return s


def main() -> None:
    # 腳本在 docs/weekly reports/其他/ 下，repo 根目錄為上三層
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent.parent
    md_path = script_dir / "代碼管制表.md"
    if not md_path.exists():
        print(f"❌ 找不到: {md_path}")
        sys.exit(1)

    text = md_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    header = "| 功能模組 | 編號 | 名稱 | 代碼 | 代碼行數 | 代碼功能描述 | 創建日期 | 最後更新日期 | 備註 |"
    sep = "|---------|------|------|------|----------|-------------|---------|-------------|------|"

    new_lines = []
    in_table = False

    for i, line in enumerate(lines):
        if line.strip().startswith("| 功能模組 |") and "代碼功能描述" in line:
            in_table = True
            new_lines.append(header)
            continue
        if in_table and line.strip().startswith("|---------|"):
            new_lines.append(sep)
            continue
        if in_table and line.strip().startswith("|"):
            raw = [x.strip() for x in line.split("|")]
            # 去掉首尾空：split('|') 得 ['', ' A ', ' B ', ..., ' Z ']，cells = raw[1:-1]
            cells = [c.strip() for c in raw[1:-1]]
            if len(cells) >= 8:
                code_idx = 3
                code_cell = cells[code_idx]
                path = extract_path(code_cell)
                full_path = repo_root / path
                num = count_lines(full_path)
                if len(cells) == 8:
                    new_cells = cells[:4] + [num] + cells[4:]
                else:
                    new_cells = cells[:4] + [num] + cells[5:]
                new_line = "| " + " | ".join(new_cells) + " |"
                new_lines.append(new_line)
            else:
                new_lines.append(line)
            continue
        if in_table and line.strip() == "":
            in_table = False
        new_lines.append(line)

    md_path.write_text("\n".join(new_lines), encoding="utf-8")
    print(f"✅ 已更新 {md_path}，已加入「代碼行數」欄位")

    # 以 openpyxl 從 MD 產出含「代碼行數」的 Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "代碼管制表"
        text = md_path.read_text(encoding="utf-8")
        for i, line in enumerate(text.splitlines()):
            if not line.strip().startswith("|") or line.strip().startswith("|---------"):
                continue
            cells = [c.strip() for c in line.split("|")][1:-1]
            if not cells:
                continue
            for j, val in enumerate(cells, 1):
                ws.cell(row=i + 1, column=j, value=val)
        for row in ws.iter_rows(min_row=1, max_row=1):
            for c in row:
                c.font = Font(bold=True)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        xlsx_dst = script_dir / "代碼管制表.xlsx"
        wb.save(xlsx_dst)
        print(f"✅ 已更新 Excel（含代碼行數）: {xlsx_dst}")
    except ImportError:
        try:
            convert_script = repo_root / "scripts" / "tools" / "convert_code_registry_to_excel.py"
            cp_md = repo_root / "scripts" / "tools" / "docs" / "系统设计文档" / "代碼管制表.md"
            cp_md.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy(md_path, cp_md)
            r = subprocess.run(
                [sys.executable, str(convert_script)],
                cwd=str(repo_root),
                capture_output=True,
                text=True,
            )
            if r.returncode == 0:
                xlsx_src = repo_root / "scripts" / "tools" / "docs" / "代碼管制表.xlsx"
                xlsx_dst = script_dir / "代碼管制表.xlsx"
                if xlsx_src.exists():
                    shutil.copy(xlsx_src, xlsx_dst)
                    print(f"✅ 已更新 Excel: {xlsx_dst}")
        except Exception as e:
            print("⚠️ 轉 Excel 跳過:", e)


if __name__ == "__main__":
    main()
