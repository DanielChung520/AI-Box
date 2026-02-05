# 代碼功能說明: 從代碼管制表 MD（含代碼行數）產出 Excel
# 創建日期: 2026-02-04
# 創建人: Daniel Chung
# 最後修改日期: 2026-02-04

"""讀取含「代碼行數」欄位的代碼管制表.md，產出對應的 .xlsx。"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    md_path = script_dir / "代碼管制表.md"
    xlsx_path = script_dir / "代碼管制表.xlsx"
    if not md_path.exists():
        print(f"❌ 找不到: {md_path}")
        sys.exit(1)

    text = md_path.read_text(encoding="utf-8")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "代碼管制表"
    row_num = 0
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|") or line.startswith("|---------"):
            continue
        parts = [p.strip() for p in line.split("|")][1:-1]
        if not parts:
            continue
        row_num += 1
        for col_num, val in enumerate(parts, 1):
            ws.cell(row=row_num, column=col_num, value=val)
    for c in ws[1]:
        c.font = Font(bold=True)
    for col in ws.columns:
        max_len = min(50, max((len(str(c.value or "")) for c in col), default=0) + 2)
        ws.column_dimensions[col[0].column_letter].width = max_len
    wb.save(xlsx_path)
    print(f"✅ 已產出 Excel（含代碼行數）: {xlsx_path}")


if __name__ == "__main__":
    main()
