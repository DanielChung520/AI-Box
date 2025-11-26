# 代碼功能說明: Excel 文件解析器
# 創建日期: 2025-01-27 23:30 (UTC+8)
# 創建人: Daniel Chung
# 最後修改日期: 2025-01-27 23:30 (UTC+8)

"""Excel 文件解析器 - 使用 openpyxl"""

from typing import Dict, Any
from .base_parser import BaseParser

try:
    from openpyxl import load_workbook

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class XlsxParser(BaseParser):
    """Excel 文件解析器"""

    def __init__(self):
        super().__init__()
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl 未安裝，請運行: pip install openpyxl")

    def parse(self, file_path: str) -> Dict[str, Any]:
        """解析 Excel 文件"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            text_parts = []
            sheets_data = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []

                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    sheet_text.append(row_text)

                sheet_full_text = "\n".join(sheet_text)
                text_parts.append(f"=== {sheet_name} ===\n{sheet_full_text}")

                sheets_data.append(
                    {
                        "name": sheet_name,
                        "rows": len(sheet_text),
                    }
                )

            full_text = "\n\n".join(text_parts)

            return {
                "text": full_text,
                "metadata": {
                    "num_sheets": len(workbook.sheetnames),
                    "sheets": sheets_data,
                    "char_count": len(full_text),
                },
            }
        except Exception as e:
            self.logger.error("Excel 文件解析失敗", file_path=file_path, error=str(e))
            raise

    def parse_from_bytes(self, file_content: bytes, **kwargs) -> Dict[str, Any]:
        """從字節內容解析 Excel"""
        try:
            from io import BytesIO

            workbook = load_workbook(BytesIO(file_content), data_only=True)
            text_parts = []
            sheets_data = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []

                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    sheet_text.append(row_text)

                sheet_full_text = "\n".join(sheet_text)
                text_parts.append(f"=== {sheet_name} ===\n{sheet_full_text}")

                sheets_data.append(
                    {
                        "name": sheet_name,
                        "rows": len(sheet_text),
                    }
                )

            full_text = "\n\n".join(text_parts)

            return {
                "text": full_text,
                "metadata": {
                    "num_sheets": len(workbook.sheetnames),
                    "sheets": sheets_data,
                    "char_count": len(full_text),
                },
            }
        except Exception as e:
            self.logger.error("Excel 解析失敗", error=str(e))
            raise

    def get_supported_extensions(self) -> list:
        return [".xlsx"]
