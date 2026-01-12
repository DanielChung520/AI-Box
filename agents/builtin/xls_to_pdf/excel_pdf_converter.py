# 代碼功能說明: Excel 到 PDF 轉換器
# 創建日期: 2026-01-11
# 創建人: Daniel Chung
# 最後修改日期: 2026-01-11

"""Excel 到 PDF 轉換器

實現通過 openpyxl + reportlab 將 Excel 轉換為 PDF。
"""

import logging
from typing import Any, Dict, Optional

try:
    from openpyxl import load_workbook
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
except ImportError:
    load_workbook = None
    SimpleDocTemplate = None

logger = logging.getLogger(__name__)


class ExcelPdfConverter:
    """Excel 到 PDF 轉換器

    提供 Excel 到 PDF 的轉換功能。
    """

    def __init__(self):
        """初始化 Excel 到 PDF 轉換器"""
        self.logger = logger

        if load_workbook is None or SimpleDocTemplate is None:
            raise ImportError(
                "Required libraries not installed. Please install: pip install openpyxl reportlab"
            )

    def convert(
        self,
        excel_path: str,
        output_path: str,
        options: Optional[Dict[str, Any]] = None,
    ) -> bool:
        """
        將 Excel 轉換為 PDF

        Args:
            excel_path: Excel 文件路徑
            output_path: 輸出 PDF 文件路徑
            options: 轉換選項

        Returns:
            轉換是否成功
        """
        try:
            # 讀取 Excel 文件
            workbook = load_workbook(excel_path, data_only=True)

            # 獲取要轉換的工作表
            worksheets = options.get("worksheets", ["all"]) if options else ["all"]
            if worksheets == "all":
                sheet_names = workbook.sheetnames
            else:
                sheet_names = [s for s in worksheets if s in workbook.sheetnames]

            if not sheet_names:
                self.logger.error("No valid worksheets found")
                return False

            # 創建 PDF 文檔
            page_size = A4
            if options and options.get("page_size") == "Letter":
                page_size = letter

            doc = SimpleDocTemplate(output_path, pagesize=page_size)

            # 構建 PDF 內容
            elements = []

            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]

                # 獲取數據範圍
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                if max_row == 0 or max_col == 0:
                    continue

                # 提取數據
                data = []
                for row in range(1, min(max_row + 1, 100)):  # 限制行數
                    row_data = []
                    for col in range(1, min(max_col + 1, 26)):  # 限制列數
                        cell = worksheet.cell(row, col)
                        row_data.append(str(cell.value) if cell.value is not None else "")
                    data.append(row_data)

                if not data:
                    continue

                # 創建表格
                table = Table(data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )

                elements.append(table)

            # 構建 PDF
            doc.build(elements)

            self.logger.info(f"PDF converted successfully: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Excel to PDF conversion error: {e}", exc_info=True)
            return False
