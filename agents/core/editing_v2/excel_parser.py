# 代碼功能說明: Excel 文件解析器
# 創建日期: 2026-01-11
# 創建人: Daniel Chung
# 最後修改日期: 2026-01-11

"""Excel 文件解析器

實現 Excel 文件的讀寫、解析和操作功能。
"""

import logging
from typing import Any, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    load_workbook = None
    Workbook = None
    Cell = None
    Worksheet = None

from agents.core.editing_v2.error_handler import EditingError, ErrorHandler

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel 文件解析器

    提供 Excel 文件的讀寫、解析和操作功能。
    """

    def __init__(self):
        """初始化 Excel 解析器"""
        self.error_handler = ErrorHandler()
        self.workbook: Optional[Any] = None
        self.file_path: Optional[str] = None

        # 檢查 openpyxl 是否可用
        if load_workbook is None:
            raise ImportError(
                "openpyxl is not installed. Please install it with: pip install openpyxl"
            )

    def load(self, file_path: str) -> None:
        """
        加載 Excel 文件

        Args:
            file_path: Excel 文件路徑

        Raises:
            EditingError: 文件不存在或格式錯誤時拋出
        """
        try:
            self.file_path = file_path
            self.workbook = load_workbook(file_path, data_only=True)
            logger.info(f"Excel file loaded: {file_path}")
        except FileNotFoundError:
            raise self.error_handler.handle_document_not_found(file_path)
        except Exception as e:
            raise self.error_handler.handle_invalid_format(f"Failed to load Excel file: {str(e)}")

    def get_worksheet(self, sheet_name: str) -> Optional[Any]:
        """
        獲取工作表

        Args:
            sheet_name: 工作表名稱

        Returns:
            工作表對象，如果不存在返回 None
        """
        if self.workbook is None:
            raise EditingError(
                code="INVALID_STATE",
                message="Workbook not loaded. Call load() first.",
            )

        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        return None

    def get_cell_value(self, sheet_name: str, cell_address: str) -> Optional[Any]:
        """
        獲取單元格值

        Args:
            sheet_name: 工作表名稱
            cell_address: 單元格地址（如 "A1" 或 "B5"）

        Returns:
            單元格值，如果不存在返回 None
        """
        worksheet = self.get_worksheet(sheet_name)
        if worksheet is None:
            return None

        try:
            cell = worksheet[cell_address]
            return cell.value
        except Exception:
            return None

    def get_range_values(self, sheet_name: str, range_address: str) -> List[List[Any]]:
        """
        獲取範圍值

        Args:
            sheet_name: 工作表名稱
            range_address: 範圍地址（如 "A1:C10"）

        Returns:
            範圍值的二維列表
        """
        worksheet = self.get_worksheet(sheet_name)
        if worksheet is None:
            return []

        try:
            cells = worksheet[range_address]
            if isinstance(cells, tuple):
                # 多行多列
                return [[cell.value for cell in row] for row in cells]
            else:
                # 單行或單列
                return [[cell.value for cell in cells]]
        except Exception as e:
            logger.warning(f"Failed to get range values: {e}")
            return []

    def set_cell_value(self, sheet_name: str, cell_address: str, value: Any) -> None:
        """
        設置單元格值

        Args:
            sheet_name: 工作表名稱
            cell_address: 單元格地址
            value: 新值
        """
        worksheet = self.get_worksheet(sheet_name)
        if worksheet is None:
            raise EditingError(
                code="TARGET_NOT_FOUND",
                message=f"Worksheet '{sheet_name}' not found",
            )

        worksheet[cell_address] = value

    def save(self, file_path: Optional[str] = None) -> None:
        """
        保存 Excel 文件

        Args:
            file_path: 保存路徑，如果為 None 則保存到原文件
        """
        if self.workbook is None:
            raise EditingError(
                code="INVALID_STATE",
                message="Workbook not loaded. Call load() first.",
            )

        save_path = file_path or self.file_path
        if save_path is None:
            raise EditingError(
                code="INVALID_STATE",
                message="No file path specified for saving.",
            )

        try:
            self.workbook.save(save_path)
            logger.info(f"Excel file saved: {save_path}")
        except Exception as e:
            raise self.error_handler.handle_invalid_format(f"Failed to save Excel file: {str(e)}")

    def get_worksheet_names(self) -> List[str]:
        """
        獲取所有工作表名稱

        Returns:
            工作表名稱列表
        """
        if self.workbook is None:
            return []
        return self.workbook.sheetnames
